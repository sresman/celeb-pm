#!/usr/bin/env python3
"""Build analysis/trigger_analysis.xlsx — a 5-sheet workbook, one sheet per trigger
type, events as rows and quarter-by-quarter (filing-to-filing) SINGLE-PERIOD returns
across columns, with allocation tracking and green/red return shading.

Values are hardcoded (no formulas) from the existing CSVs — no recalc needed.
Every return cell is one filing-to-filing period return; nothing is compounded.

Usage:
    python tools/transcripts/build_trigger_workbook.py
"""

from __future__ import annotations

import datetime as dt
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

try:  # runs both as a direct script and under `-m`
    from . import generate_13f_triggers as trig
except ImportError:
    import generate_13f_triggers as trig  # type: ignore[import-not-found, no-redef]

ANALYSIS_DIR = trig.ANALYSIS_DIR
RECLASS_PATH = ANALYSIS_DIR / "ai_basket_reclassification.json"
TRIGGERS_IN = ANALYSIS_DIR / "13f_signal_triggers_clean.csv"
UNIVERSAL_IN = ANALYSIS_DIR / "filing_to_filing_returns_universal.csv"
OUT_PATH = ANALYSIS_DIR / "trigger_analysis.xlsx"

FWD_QUARTERS = 24  # Q+1 .. Q+24 for the variable-length sheets
SENTINELS = {"PRE_IPO", "NO_DATA", ""}
RAMP_EXCLUDED_BUCKETS = {"AI/Hyperscaler", "AI/EV"}  # narrow "picks-and-shovels" basket

HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center")
GREEN = PatternFill("solid", fgColor="DCF0DC")  # (220,240,220)
RED = PatternFill("solid", fgColor="F0DCDC")    # (240,220,220)
RET_FMT = "+0.0%;-0.0%;0.0%"
ALLOC_FMT = "0.0%"
DATE_FMT = "yyyy-mm-dd"


# --------------------------------------------------------------------------
# Lookups
# --------------------------------------------------------------------------


def build_lookups() -> tuple[
    dict[tuple[str, str], float],       # return: (ticker, period_start) -> pct
    dict[tuple[str, str], float],       # ticker alloc: (ticker, filing_date) -> weight
    dict[tuple[str, str], float],       # subtheme alloc: (subtheme, filing_date) -> weight
    list[str],                          # sorted filing dates (25)
]:
    ret: dict[tuple[str, str], float] = {}
    filings_set: set[str] = set()
    for r in trig._read_csv(UNIVERSAL_IN):
        filings_set.add(r["period_start_filing"])
        filings_set.add(r["period_end_filing"])
        val = r["return_pct"]
        if val not in SENTINELS:
            ret[(r["ticker"], r["period_start_filing"])] = float(val)

    reclass = json.loads(RECLASS_PATH.read_text(encoding="utf-8"))
    views = trig.find_views_dir()
    ticker_alloc: dict[tuple[str, str], float] = {}
    subtheme_alloc: dict[tuple[str, str], float] = {}
    for r in trig._read_csv(views / "position_lifecycles.csv"):
        if r["security_type"] != "COMMON" or r["change_type"] == "EXIT":
            continue
        weight = trig._f(r["weight_pct"]) or 0.0
        fd = r["filing_date"]
        ticker_alloc[(r["ticker"], fd)] = weight
        is_ai, subtheme = trig.resolve_ai(reclass, r["ticker"], fd, r["theme"])
        if is_ai and subtheme is not None:
            key = (subtheme, fd)
            subtheme_alloc[key] = subtheme_alloc.get(key, 0.0) + weight

    return ret, ticker_alloc, subtheme_alloc, sorted(filings_set)


def basket_return(ret: dict[tuple[str, str], float], tickers: list[str], start: str) -> float | None:
    """Equal-weight average of constituents' single-period returns; drop missing."""
    vals = [ret[(t, start)] for t in tickers if (t, start) in ret]
    return sum(vals) / len(vals) if vals else None


def cw_return(ret: dict[tuple[str, str], float], pairs: list[tuple[str, float]], start: str) -> float | None:
    """Capital-weighted single-period return; renormalize among constituents with
    data that period (denominator = summed weights of the available names)."""
    num = den = 0.0
    for ticker, weight in pairs:
        v = ret.get((ticker, start))
        if v is not None:
            num += weight * v
            den += weight
    return num / den if den > 0 else None


def build_ramp_holdings() -> dict[str, list[tuple[str, float]]]:
    """Per filing date: the AI picks-and-shovels COMMON non-EXIT holdings
    (ticker, weight_pct), sorted by weight desc. Same narrow filter as Trigger 1."""
    reclass = json.loads(RECLASS_PATH.read_text(encoding="utf-8"))
    views = trig.find_views_dir()
    holdings: dict[str, list[tuple[str, float]]] = {}
    for r in trig._read_csv(views / "position_lifecycles.csv"):
        if r["security_type"] != "COMMON" or r["change_type"] == "EXIT":
            continue
        is_ai, bucket = trig.resolve_ai(reclass, r["ticker"], r["filing_date"], r["theme"])
        if is_ai and bucket not in RAMP_EXCLUDED_BUCKETS:
            holdings.setdefault(r["filing_date"], []).append(
                (r["ticker"], trig._f(r["weight_pct"]) or 0.0))
    for fd in holdings:
        holdings[fd].sort(key=lambda x: x[1], reverse=True)
    return holdings


# --------------------------------------------------------------------------
# Cell writers
# --------------------------------------------------------------------------


def _put_ret(ws: Worksheet, row: int, col: int, pct: float | None) -> None:
    if pct is None:
        return
    cell = ws.cell(row, col, pct / 100.0)
    cell.number_format = RET_FMT
    if pct > 0:
        cell.fill = GREEN
    elif pct < 0:
        cell.fill = RED


def _put_weight(ws: Worksheet, row: int, col: int, pct: float | None, fmt: str = ALLOC_FMT) -> None:
    if pct is None:
        return
    cell = ws.cell(row, col, pct / 100.0)
    cell.number_format = fmt


def _put_date(ws: Worksheet, row: int, col: int, iso: str) -> None:
    cell = ws.cell(row, col, dt.date.fromisoformat(iso))
    cell.number_format = DATE_FMT


def _finish(ws: Worksheet, headers: list[str], id_widths: list[float], period_w: float = 10) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for i, w in enumerate(id_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(id_widths) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = period_w
    ws.freeze_panes = f"{get_column_letter(len(id_widths) + 1)}2"


def _tickers(cell: str) -> list[str]:
    return [t.strip() for t in cell.split(",") if t.strip()]


# Buying-detail block shared by the Ramp and RampBasket sheets (net-buying trigger metrics
# + what was bought/sold/entered/exited). Placed after the weight columns, before returns.
BUYING_DETAIL_HEADERS = [
    "net_buying_pct", "net_buying_dollars", "gross_buying_pct", "gross_selling_pct",
    "tickers_bought", "tickers_sold", "tickers_new", "tickers_exited",
]
BUYING_DETAIL_WIDTHS: list[float] = [14, 16, 14, 14, 30, 30, 22, 22]


def _put_buying_detail(ws: Worksheet, row: int, col: int, e: dict[str, str]) -> int:
    """Write the 8 buying-detail cells starting at `col`; return the next free column."""
    _put_weight(ws, row, col, trig._f(e.get("net_buying_pct")))
    nbd = trig._f(e.get("net_buying_dollars"))
    if nbd is not None:
        ws.cell(row, col + 1, nbd).number_format = "#,##0"
    _put_weight(ws, row, col + 2, trig._f(e.get("gross_buying_pct")))
    _put_weight(ws, row, col + 3, trig._f(e.get("gross_selling_pct")))
    ws.cell(row, col + 4, e.get("tickers_bought", ""))
    ws.cell(row, col + 5, e.get("tickers_sold", ""))
    ws.cell(row, col + 6, e.get("tickers_new", ""))
    ws.cell(row, col + 7, e.get("tickers_exited", ""))
    return col + 8


# --------------------------------------------------------------------------
# Sheets
# --------------------------------------------------------------------------


def sheet_ramp(ws: Worksheet, events: list[dict[str, str]], ret: dict[tuple[str, str], float],
               filings: list[str], fidx: dict[str, int]) -> None:
    pre = [4, 3, 2, 1]
    fwd = list(range(1, 9))
    headers = (["filing_date", "ai_weight_change", "ai_weight_current"] + BUYING_DETAIL_HEADERS
               + [f"Q-{i}_SMH" for i in pre] + [f"Q+{j}_SMH" for j in fwd])
    _finish(ws, headers, [12.0, 16.0, 16.0, *BUYING_DETAIL_WIDTHS])
    for row, e in enumerate(events, 2):
        k = fidx[e["filing_date"]]
        _put_date(ws, row, 1, e["filing_date"])
        _put_weight(ws, row, 2, trig._f(e["ai_weight_change"]), RET_FMT)
        _put_weight(ws, row, 3, trig._f(e["ai_weight_current"]))
        col = _put_buying_detail(ws, row, 4, e)  # cols 4..11
        for i in pre:
            if k - i >= 0:
                _put_ret(ws, row, col, ret.get(("SMH", filings[k - i])))
            col += 1
        for j in fwd:
            if k + j <= len(filings) - 1:
                _put_ret(ws, row, col, ret.get(("SMH", filings[k + j - 1])))
            col += 1


def sheet_ramp_basket(ws: Worksheet, events: list[dict[str, str]], ret: dict[tuple[str, str], float],
                      holdings: dict[str, list[tuple[str, float]]], filings: list[str],
                      fidx: dict[str, int]) -> None:
    pre = [4, 3, 2, 1]
    fwd = list(range(1, 9))
    max_n = max((len(holdings.get(e["filing_date"], [])) for e in events), default=0)
    headers = ["filing_date", "ai_weight_change", "ai_weight_current"] + BUYING_DETAIL_HEADERS
    for i in range(1, max_n + 1):
        headers += [f"ticker_{i}", f"wt_{i}"]
    for i in pre:
        headers += [f"Q-{i}_EW", f"Q-{i}_CW", f"Q-{i}_SMH"]
    for j in fwd:
        headers += [f"Q+{j}_EW", f"Q+{j}_CW", f"Q+{j}_SMH"]
    _finish(ws, headers, [12.0, 16.0, 16.0, *BUYING_DETAIL_WIDTHS])

    comp_start = 4 + len(BUYING_DETAIL_HEADERS)  # composition begins after the detail block
    for row, e in enumerate(events, 2):
        k = fidx[e["filing_date"]]
        pairs = holdings.get(e["filing_date"], [])
        tickers = [t for t, _ in pairs]
        _put_date(ws, row, 1, e["filing_date"])
        _put_weight(ws, row, 2, trig._f(e["ai_weight_change"]), RET_FMT)
        _put_weight(ws, row, 3, trig._f(e["ai_weight_current"]))
        _put_buying_detail(ws, row, 4, e)  # cols 4..11
        col = comp_start
        for ticker, weight in pairs:
            ws.cell(row, col, ticker)
            _put_weight(ws, row, col + 1, weight)
            col += 2
        col = comp_start + max_n * 2  # start of the period block (past all ticker/wt pairs)
        starts = ([filings[k - i] if k - i >= 0 else None for i in pre]
                  + [filings[k + j - 1] if k + j <= len(filings) - 1 else None for j in fwd])
        for start in starts:
            if start is not None:
                _put_ret(ws, row, col, basket_return(ret, tickers, start))
                _put_ret(ws, row, col + 1, cw_return(ret, pairs, start))
                _put_ret(ws, row, col + 2, ret.get(("SMH", start)))
            col += 3


def sheet_new_subtheme(ws: Worksheet, events: list[dict[str, str]], ret: dict[tuple[str, str], float],
                       subtheme_alloc: dict[tuple[str, str], float], filings: list[str],
                       fidx: dict[str, int]) -> None:
    headers = ["filing_date", "subtheme", "entering_tickers", "total_subtheme_weight"]
    for j in range(1, FWD_QUARTERS + 1):
        headers += [f"Q+{j}_alloc", f"Q+{j}_basket", f"Q+{j}_SMH"]
    _finish(ws, headers, [12, 26, 26, 18])
    for row, e in enumerate(events, 2):
        k = fidx[e["filing_date"]]
        basket = _tickers(e["entering_tickers"])
        _put_date(ws, row, 1, e["filing_date"])
        ws.cell(row, 2, e["subtheme"])
        ws.cell(row, 3, e["entering_tickers"])
        _put_weight(ws, row, 4, trig._f(e["total_subtheme_weight"]))
        col = 5
        for j in range(1, FWD_QUARTERS + 1):
            if k + j <= len(filings) - 1:
                _put_weight(ws, row, col, subtheme_alloc.get((e["subtheme"], filings[k + j])))
                _put_ret(ws, row, col + 1, basket_return(ret, basket, filings[k + j - 1]))
                _put_ret(ws, row, col + 2, ret.get(("SMH", filings[k + j - 1])))
            col += 3


def sheet_new_position(ws: Worksheet, events: list[dict[str, str]], ret: dict[tuple[str, str], float],
                       ticker_alloc: dict[tuple[str, str], float], filings: list[str],
                       fidx: dict[str, int]) -> None:
    headers = ["filing_date", "ticker", "theme", "initial_weight_pct"]
    for j in range(1, FWD_QUARTERS + 1):
        headers += [f"Q+{j}_alloc", f"Q+{j}_ticker", f"Q+{j}_SMH"]
    _finish(ws, headers, [12, 10, 24, 16])
    for row, e in enumerate(events, 2):
        k = fidx[e["filing_date"]]
        ticker = e["ticker"]
        _put_date(ws, row, 1, e["filing_date"])
        ws.cell(row, 2, ticker)
        ws.cell(row, 3, e["theme"])
        _put_weight(ws, row, 4, trig._f(e["initial_weight_pct"]))
        col = 5
        for j in range(1, FWD_QUARTERS + 1):
            if k + j <= len(filings) - 1:
                _put_weight(ws, row, col, ticker_alloc.get((ticker, filings[k + j])))
                _put_ret(ws, row, col + 1, ret.get((ticker, filings[k + j - 1])))
                _put_ret(ws, row, col + 2, ret.get(("SMH", filings[k + j - 1])))
            col += 3


def sheet_cross(ws: Worksheet, events: list[dict[str, str]], ret: dict[tuple[str, str], float],
                subtheme_alloc: dict[tuple[str, str], float], filings: list[str],
                fidx: dict[str, int]) -> None:
    pre = [4, 3, 2, 1]
    headers = ["filing_date", "subtheme", "subtheme_weight_prior",
               "subtheme_weight_current", "active_tickers", "all_tickers"]
    for i in pre:
        headers += [f"Q-{i}_basket", f"Q-{i}_SMH"]
    for j in range(1, FWD_QUARTERS + 1):
        headers += [f"Q+{j}_alloc", f"Q+{j}_basket", f"Q+{j}_SMH"]
    _finish(ws, headers, [12, 26, 20, 20, 28, 34])
    for row, e in enumerate(events, 2):
        k = fidx[e["filing_date"]]
        basket = _tickers(e["all_tickers_in_subtheme"])
        _put_date(ws, row, 1, e["filing_date"])
        ws.cell(row, 2, e["subtheme"])
        _put_weight(ws, row, 3, trig._f(e["subtheme_weight_prior"]))
        _put_weight(ws, row, 4, trig._f(e["subtheme_weight_current"]))
        ws.cell(row, 5, e["active_tickers"])
        ws.cell(row, 6, e["all_tickers_in_subtheme"])
        col = 7
        for i in pre:
            if k - i >= 0:
                _put_ret(ws, row, col, basket_return(ret, basket, filings[k - i]))
                _put_ret(ws, row, col + 1, ret.get(("SMH", filings[k - i])))
            col += 2
        for j in range(1, FWD_QUARTERS + 1):
            if k + j <= len(filings) - 1:
                _put_weight(ws, row, col, subtheme_alloc.get((e["subtheme"], filings[k + j])))
                _put_ret(ws, row, col + 1, basket_return(ret, basket, filings[k + j - 1]))
                _put_ret(ws, row, col + 2, ret.get(("SMH", filings[k + j - 1])))
            col += 3


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------


def _by_type(rows: list[dict[str, str]], ttype: str) -> list[dict[str, str]]:
    return sorted((r for r in rows if r["trigger_type"] == ttype),
                  key=lambda r: r["filing_date"])


def main() -> int:
    ret, ticker_alloc, subtheme_alloc, filings = build_lookups()
    fidx = {d: i for i, d in enumerate(filings)}
    rows = trig._read_csv(TRIGGERS_IN)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    sheet_ramp(wb.create_sheet("Ramp"),
               _by_type(rows, "AI_BASKET_RAMP"), ret, filings, fidx)
    sheet_new_subtheme(wb.create_sheet("NewSubtheme"),
                       _by_type(rows, "NEW_AI_SUBTHEME"), ret, subtheme_alloc, filings, fidx)
    sheet_new_position(wb.create_sheet("NewPosition"),
                       _by_type(rows, "NEW_AI_POSITION_2PCT"), ret, ticker_alloc, filings, fidx)
    sheet_cross(wb.create_sheet("Cross4pct"),
                _by_type(rows, "AI_SUBTHEME_ACTIVE_CROSS_4PCT"), ret, subtheme_alloc, filings, fidx)
    sheet_cross(wb.create_sheet("Cross2pct"),
                _by_type(rows, "AI_SUBTHEME_ACTIVE_CROSS_2PCT"), ret, subtheme_alloc, filings, fidx)
    sheet_ramp_basket(wb.create_sheet("RampBasket"),
                      _by_type(rows, "AI_BASKET_RAMP"), ret, build_ramp_holdings(), filings, fidx)

    wb.save(OUT_PATH)
    counts = {ws.title: ws.max_row - 1 for ws in wb.worksheets}
    print(f"Wrote {OUT_PATH.relative_to(trig.REPO_ROOT)}")
    for title, n in counts.items():
        print(f"  {title}: {n} events")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
