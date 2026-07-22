"""Build the repeat-mention signal-events sheet for the noise-vs-criteria test.

Restructures the event grain from theme_returns_v2's "one row per criterion-triggered
event" to **one row per (theme, unique-mention-date)**, so every 2nd+ mention of a
theme is represented regardless of whether it meets an existing signal criterion.

Each row carries:
  - mention_number / total_theme_mentions (per-theme unique-date ordinal)
  - is_repeat_mention  = mention_number >= 2   (the broad superset population)
  - the existing signal criteria as SEPARATE booleans:
      is_third_within_1yr, is_first_hc_not_first_mention,
      is_first_with_tickers, is_hc_high_profile_venue
    plus is_first_mention / is_first_mention_hc / is_thesis_reversal for context
  - meets_existing_criteria = OR of the 4 operator-listed criteria above
  - is_derisk_signal / is_thesis_close / exclude_from_long_stats (set via overrides)
  - returns at 7 horizons (1m/1q/6m/9m/1y/18m/2y) vs SMH and SPY benchmarks

Applies the durable manual-override layer (analysis/manual_overrides.json) AFTER
clustering + basket resolution, immediately before returns — cluster overrides
re-theme theses; event overrides set basket/source/direction/flags. Overrides win.

Three-way slice (repeat mentions, excluding thesis-close/reversal rows):
  1. population  = is_repeat_mention
  2. signal      = is_repeat_mention & meets_existing_criteria
  3. control     = is_repeat_mention & ~meets_existing_criteria

Usage:
    python -m tools.transcripts.build_repeat_mention_events [--force-refetch]
"""

from __future__ import annotations

import argparse
import bisect
import csv
import json
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]

from . import targets
from .theme_returns_v2 import (
    CALENDAR_SYMBOL,
    INSUFFICIENT_DATA,
    NO_BASKET,
    NO_DATA,
    PAIR_TRADE,
    SMH_SYMBOL,
    _days_between,
    _is_hc,
    _ticker_return,
    _venue_match,
    apply_cluster_overrides,
    apply_event_override,
    cluster_theses,
    fetch_prices,
    find_event_override,
    is_stance_reversal,
    load_api_key,
    load_overrides,
    resolve_basket,
)

ANALYSIS_DIR = targets.REPO_ROOT / "analysis"
TIMELINE_JSON = ANALYSIS_DIR / "thesis_timeline_v2_flat.json"
BASKETS_JSON = ANALYSIS_DIR / "theme_baskets_v3.json"
OUTPUT_CSV = ANALYSIS_DIR / "step4_signal_events_v7_with_returns_extended.csv"
OUTPUT_XLSX = ANALYSIS_DIR / "step4_signal_events_v7_with_returns_extended.xlsx"

HORIZONS: dict[str, int] = {
    "1m": 21, "1q": 63, "6m": 126, "9m": 189, "1y": 252, "18m": 378, "2y": 504,
}
SPY_SYMBOL = "SPY"

META_FIELDS = [
    "date", "theme", "mention_number", "total_theme_mentions",
    "is_repeat_mention", "meets_existing_criteria",
    "is_first_mention", "is_first_mention_hc", "is_first_hc_not_first_mention",
    "is_third_within_1yr", "is_first_with_tickers", "is_hc_high_profile_venue",
    "is_thesis_reversal", "is_thesis_close", "is_derisk_signal",
    "exclude_from_long_stats", "criteria_met",
    "confidence", "source", "summary", "summary_extended", "tickers_direct",
    "resolved_basket", "basket_source", "basket_direction", "basket_asterisk",
    "override_note",
]
RETURN_FIELDS = (
    [f"ret_{h}" for h in HORIZONS]
    + [f"smh_{h}" for h in HORIZONS]
    + [f"excess_{h}" for h in HORIZONS]
    + [f"spy_{h}" for h in HORIZONS]
    + [f"excess_spy_{h}" for h in HORIZONS]
)
OUTPUT_FIELDS = META_FIELDS + RETURN_FIELDS

SIGNAL_CRITERIA = (
    "is_third_within_1yr", "is_first_hc_not_first_mention",
    "is_first_with_tickers", "is_hc_high_profile_venue",
)
BOOL_DEFAULTS = {
    "is_thesis_reversal": "FALSE", "is_thesis_close": "FALSE",
    "is_derisk_signal": "FALSE", "exclude_from_long_stats": "FALSE",
    "override_note": "",
}


def _b(x: bool) -> str:
    return "TRUE" if x else "FALSE"


def generate_mention_rows(
    theme: str, theses: list[dict[str, Any]], overrides: dict[str, Any]
) -> list[dict[str, Any]]:
    """One row per unique mention-date for a theme, with criteria booleans."""
    if not theses:
        return []
    theses = sorted(theses, key=lambda t: (t["date"], t["thesis_id"]))
    unique_dates = sorted({t["date"] for t in theses})
    total_mentions = len(unique_dates)
    date_index = {d: i + 1 for i, d in enumerate(unique_dates)}
    by_date: dict[str, list[dict[str, Any]]] = {}
    for t in theses:
        by_date.setdefault(t["date"], []).append(t)

    first_date = unique_dates[0]
    earliest_hc_date = next(
        (d for d in unique_dates if any(_is_hc(t) for t in by_date[d])), None
    )
    earliest_ticker_date = next(
        (t["date"] for t in theses if t.get("tickers_direct")), None
    )

    rows: list[dict[str, Any]] = []
    for d in unique_dates:
        day_theses = by_date[d]
        # override-aware rep selection: prefer a thesis matched by a thesis-specific
        # (summary_contains) override, else the HC one, else the first.
        rep = next(
            (t for t in day_theses if find_event_override(
                overrides, theme=theme, date=d, summary=str(t.get("summary", "")),
                require_summary=True) is not None),
            None,
        )
        if rep is None:
            rep = next((t for t in day_theses if _is_hc(t)), day_theses[0])
        mnum = date_index[d]

        is_first_mention = d == first_date
        is_first_mention_hc = is_first_mention and any(_is_hc(t) for t in day_theses)
        is_first_hc_not_first = (d == earliest_hc_date) and (d != first_date)
        is_third_1yr = (mnum == 3) and _days_between(first_date, d) <= 365
        is_first_tickers = (earliest_ticker_date is not None) and (d == earliest_ticker_date)
        is_hc_venue = any(
            _is_hc(t) and _venue_match(str(t.get("source", ""))) for t in day_theses
        )
        is_reversal = any(
            is_stance_reversal(str(t.get("summary", ""))) for t in day_theses
        )

        flags = {
            "is_first_mention": is_first_mention,
            "is_first_mention_hc": is_first_mention_hc,
            "is_first_hc_not_first_mention": is_first_hc_not_first,
            "is_third_within_1yr": is_third_1yr,
            "is_first_with_tickers": is_first_tickers,
            "is_hc_high_profile_venue": is_hc_venue,
            "is_thesis_reversal": is_reversal,
        }
        meets = any(flags[c] for c in SIGNAL_CRITERIA)
        criteria_met = ", ".join(c[3:] for c in SIGNAL_CRITERIA if flags[c])

        row: dict[str, Any] = {
            "date": d, "theme": theme,
            "mention_number": mnum, "total_theme_mentions": total_mentions,
            "is_repeat_mention": _b(mnum >= 2),
            "meets_existing_criteria": _b(meets),
            **{k: _b(v) for k, v in flags.items()},
            **BOOL_DEFAULTS,
            "is_thesis_reversal": _b(is_reversal),
            "criteria_met": criteria_met,
            "confidence": rep.get("confidence", ""),
            "source": rep.get("source", ""),
            "summary": rep.get("summary", ""),
            "summary_extended": rep.get("summary_extended", ""),
            "tickers_direct": ", ".join(rep.get("tickers_direct", []) or []),
        }
        rows.append(row)
    return rows


def compute_returns_7h(
    event_date: str, basket: list[str], direction: str, skip_label: str | None,
    prices: dict[str, dict[str, float]], calendar: list[str],
) -> dict[str, Any]:
    """Basket (EW, SHORT inverts) + SMH + SPY benchmarks + excess, at 7 horizons."""
    smh = prices.get(SMH_SYMBOL, {})
    spy = prices.get(SPY_SYMBOL, {})
    out: dict[str, Any] = {}
    start_idx = bisect.bisect_left(calendar, event_date)

    for label, n in HORIZONS.items():
        if start_idx >= len(calendar) or start_idx + n >= len(calendar):
            for fam in ("ret", "smh", "excess", "spy", "excess_spy"):
                out[f"{fam}_{label}"] = INSUFFICIENT_DATA
            continue
        sd, ed = calendar[start_idx], calendar[start_idx + n]
        smh_ret = _ticker_return(smh, sd, ed)
        spy_ret = _ticker_return(spy, sd, ed)
        out[f"smh_{label}"] = round(smh_ret, 2) if smh_ret is not None else INSUFFICIENT_DATA
        out[f"spy_{label}"] = round(spy_ret, 2) if spy_ret is not None else INSUFFICIENT_DATA

        if skip_label is not None:  # NO_BASKET / PAIR_TRADE
            out[f"ret_{label}"] = skip_label
            out[f"excess_{label}"] = skip_label
            out[f"excess_spy_{label}"] = skip_label
            continue
        per = [
            r for t in basket
            if (r := _ticker_return(prices.get(t, {}), sd, ed)) is not None
        ]
        if not per:
            out[f"ret_{label}"] = NO_DATA
            out[f"excess_{label}"] = NO_DATA
            out[f"excess_spy_{label}"] = NO_DATA
            continue
        raw = sum(per) / len(per)
        signed = -raw if direction == "SHORT" else raw
        out[f"ret_{label}"] = round(signed, 2)
        out[f"excess_{label}"] = (
            round(signed - smh_ret, 2) if smh_ret is not None else INSUFFICIENT_DATA
        )
        out[f"excess_spy_{label}"] = (
            round(signed - spy_ret, 2) if spy_ret is not None else INSUFFICIENT_DATA
        )
    return out


def _slice_stats(rows: list[dict[str, Any]]) -> dict[str, Any]:
    stats: dict[str, Any] = {"n": len(rows)}
    for label in HORIZONS:
        for fam in ("ret", "excess", "excess_spy"):
            vals = [r[f"{fam}_{label}"] for r in rows
                    if isinstance(r[f"{fam}_{label}"], (int, float))]
            stats[f"{fam}_{label}"] = round(sum(vals) / len(vals), 2) if vals else None
        rvals = [r[f"ret_{label}"] for r in rows
                 if isinstance(r[f"ret_{label}"], (int, float))]
        stats[f"winrate_{label}"] = (
            round(100 * sum(1 for v in rvals if v > 0) / len(rvals), 1) if rvals else None
        )
    return stats


def _slice_groups(rows: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    # repeat mentions, excluding thesis-close / reversal rows from long scoring
    repeats = [r for r in rows
               if r["is_repeat_mention"] == "TRUE" and r["exclude_from_long_stats"] != "TRUE"]
    g2 = [r for r in repeats if r["meets_existing_criteria"] == "TRUE"]
    g3 = [r for r in repeats if r["meets_existing_criteria"] == "FALSE"]
    return [("1_all_repeat_mentions", repeats),
            ("2_signal_meets_criteria", g2),
            ("3_control_no_criteria", g3)]


def write_xlsx(rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "signal_events"
    ws.append(OUTPUT_FIELDS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(f, "") for f in OUTPUT_FIELDS])
    ws.freeze_panes = "A2"

    ss = wb.create_sheet("slice_summary")
    cols = ["group", "n"] + [
        f"{fam}_{h}" for h in HORIZONS for fam in ("ret", "excess", "excess_spy", "winrate")
    ]
    ss.append(cols)
    for c in ss[1]:
        c.font = Font(bold=True)
    for name, grp in _slice_groups(rows):
        st = _slice_stats(grp)
        ss.append([name] + [st.get(c, "") for c in cols[1:]])
    ss.freeze_panes = "B2"
    wb.save(OUTPUT_XLSX)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--force-refetch", action="store_true")
    args = parser.parse_args()

    api_key = load_api_key()
    baskets = json.loads(BASKETS_JSON.read_text())
    theses = json.loads(TIMELINE_JSON.read_text())
    overrides = load_overrides()
    print(f"Loaded {len(theses)} theses, {len(baskets)} themes, "
          f"{len(overrides['cluster_overrides'])} cluster + "
          f"{len(overrides['event_overrides'])} event overrides.")

    assignments, _ = cluster_theses(theses, baskets)
    apply_cluster_overrides(assignments, overrides)

    rows: list[dict[str, Any]] = []
    for theme, theme_theses in assignments.items():
        rows.extend(generate_mention_rows(theme, theme_theses, overrides))
    print(f"Generated {len(rows)} mention rows "
          f"({sum(1 for r in rows if r['is_repeat_mention'] == 'TRUE')} repeat mentions).")

    prices = fetch_prices(api_key, args.force_refetch)
    calendar = sorted(prices.get(CALENDAR_SYMBOL, {}).keys())
    if not calendar:
        raise SystemExit("No SPY calendar — cannot anchor returns.")

    for r in rows:
        tickers, source, direction, asterisk = resolve_basket(r["theme"], r["date"], baskets)
        r["resolved_basket"] = ", ".join(tickers)
        r["basket_source"] = source if source else NO_BASKET
        r["basket_direction"] = direction
        r["basket_asterisk"] = str(asterisk)
        ov = find_event_override(
            overrides, theme=r["theme"], date=r["date"],
            summary=str(r.get("summary", "")), mention_number=r["mention_number"],
        )
        if ov is not None:
            apply_event_override(r, ov)
        basket = [x.strip() for x in r["resolved_basket"].split(",") if x.strip()]
        skip = None
        if r["basket_source"] == PAIR_TRADE or r["basket_direction"] == PAIR_TRADE:
            skip = PAIR_TRADE
        elif not basket:
            skip = NO_BASKET
        r.update(compute_returns_7h(r["date"], basket, r["basket_direction"], skip, prices, calendar))

    rows.sort(key=lambda e: (e["date"], e["theme"]))
    with OUTPUT_CSV.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {len(rows)} rows → {OUTPUT_CSV}")

    write_xlsx(rows)
    print(f"Wrote xlsx → {OUTPUT_XLSX}")

    for name, grp in _slice_groups(rows):
        st = _slice_stats(grp)
        print(f"  {name:<26} n={st['n']:<4} ret_1y={st.get('ret_1y')} "
              f"excess_1y={st.get('excess_1y')} winrate_1y={st.get('winrate_1y')}")


if __name__ == "__main__":
    main()
